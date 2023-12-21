from pathlib import Path
import os
from traceback import format_exc
from datetime import datetime
import subprocess, sys
from sys import exit 
import json 

import tkinter
import tkinter.messagebox
import argparse 

import xml.etree.ElementTree as ET
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Alignment

def process_results(node, results):
    if node.tag == 'test-run':
        #make sure tests were run
        tests_run = node.attrib['testcasecount']
        if tests_run == 0:
            return
        
    if node.tag == 'test-case':
        #log down the row
        test_name = node.attrib['name']
        results[test_name] = node.attrib
        
        #find the output and add it to it. also, ensure that each test HAS output
        output = find_output(node)
        if output is not None:
            results[test_name]['output'] = output.text       
    else:
        for child in node:
            process_results(child, results)

#find the output that immediately following a test case
def find_output(node):
    if node.tag == 'output':
        return node
    for child in node:
        output = find_output(child)
        if output is not None:
            return output

#converts results into spreadsheet and formats 
def format_results(results, results_path, results_name):
    df_results = pd.DataFrame(results)
    df_results = df_results.transpose()
    
    #drop redundant/unused attributes
    drop_cols = ['name', 'methodname', 'fullname', 'classname',
            'id', 'runstate', 'start-time', 'end-time','label']
    result_cols = df_results.columns
    for col in result_cols:
        if col in drop_cols:
            df_results = df_results.drop([col], axis=1)
    
    #ensure that columns are in the same order
    col_order = {}
    for col_num, col_name in enumerate(['result', 'duration', 'output', 'asserts', 'seed']):
        col_order[col_name] = col_num

    #swap until they orders match
    col_list = list(df_results.columns)
    for i in range(len(col_order)):
        num_swaps = 0
        for j in range(len(col_order)-1):
            curr_column = col_list[j]
            next_column = col_list[j+1]
            if col_order[curr_column] > col_order[next_column]:
                x, y = col_list.index(curr_column), col_list.index(next_column)
                col_list[y], col_list[x] = col_list[x], col_list[y]
                df_results = df_results[col_list]
                num_swaps += 1
        if not num_swaps:
            break

    #make it a pretty excel spreadsheet
    try:
        xlsx_path = str(results_path).replace('results.xml', f'{results_name}')
        xlsx_path = Path(xlsx_path)
        df_results.to_excel(xlsx_path)
        wb = openpyxl.load_workbook(xlsx_path)
        ws = wb['Sheet1']
        
        #fill in the pass/fail colors to pretty it up
        fill_colors = {}
        fill_colors['Passed'] = PatternFill(patternType='solid', fgColor='00A4FF72')        #green
        fill_colors['Failed'] = PatternFill(patternType='solid', fgColor='00FF7A72')        #red
        fill_colors['Skipped'] = PatternFill(patternType='solid', fgColor='0070C1FF')       #blue
        fill_colors['Inconclusive'] = PatternFill(patternType='solid', fgColor='00BABABA')  #grey
        for i in range(len(df_results.index)):
            row_num = i + 2
            result_cell = f'B{row_num}'
            ws[result_cell].fill = fill_colors[ws[result_cell].internal_value]

            output_cell = f'D{row_num}'
            if len(ws[output_cell].value[-1]) > 0 and ws[output_cell].value[-1] == '\n':
                #by default there's an extra space at the end of the output; remove it
                ws[output_cell].value = ws[output_cell].value[:-1]

                #split by line and find start/end of output
                output_lines = ws[output_cell].value.split('\n')
                start_line = 0
                end_line = len(output_lines) - 1
                for i in range(len(output_lines)):
                    if 'Test start' in output_lines[i]:
                        start_line = i
                    if 'Test end' in output_lines[i]:
                        end_line = i
                        break
                
                #get the stuff between the boundaries and write to it
                ws[output_cell].value = '\n'.join(output_lines[start_line:end_line+1])

        #also resize and align cells vertically
        for column_cells in ws.columns:
            length = max(len(str(cell.value)) for cell in column_cells) + 2 
            ws.column_dimensions[column_cells[0].column_letter].width = length
            for cell in column_cells:
                cell.alignment = Alignment(vertical='top', wrap_text=True)

        wb.save(xlsx_path)

    except PermissionError:
        exit('Could not open results.xml. If you have it open, please close it and try again.')

#runs command and returns decoded stdout. splits if delim is anything but empty string
def run_cmd(args: list[str], delim='\r\n', return_line=None, raw=False) -> list[str]:
    subprocess_results = subprocess.run(args, stdout=subprocess.PIPE, stderr=subprocess.STDOUT, stdin=subprocess.DEVNULL)
    if raw: return subprocess_results
    subprocess_output = subprocess_results.stdout.decode()
    if len(delim) > 0: 
        subprocess_output = subprocess_output.split(delim)
        if return_line is not None:
            subprocess_output = subprocess_output[return_line]
    return subprocess_output

#checks that the svn command line tool is installed
def check_svn_clt():
    print('Checking that svn command line tool is installed.', flush=True)
    try:
        run_cmd(['svn', '--version'])
    except FileNotFoundError:
        warning = 'The hook script requires the svn command line tool to run properly. To run the tests, please install it. The commit will continue without running any tests.'
        if args.show_ui:
            tkinter.messagebox.showwarning('Warning!', warning)
        else:
            print(warning)
        exit(1) #return normally

#test runner hook script. should be run before commits
def run_tests():
    #check if executable is already running
    print('Checking that hook script isn\'t already running.', flush=True)
    exec_name = os.path.basename(sys.executable)
    tasklist_output = run_cmd(['tasklist'], delim='')
    if tasklist_output.count(exec_name) > 2: #for some reason there's two processes? 
        exit('Hook script is already running!')

    #check that the unity exec path is valid, and version is correct
    print('Checking unity editor\'s path, version.')
    try:
        unity_version = run_cmd([editor_path, '--version'], return_line=0)
        if editor_version not in unity_version:
            raise FileNotFoundError
        
        print('Valid editor found!')
        unity_path = editor_path
    except FileNotFoundError:
        #invalid unity path; find the executable, if any
        print('Editor path is invalid. Looking for editor.')
        possible_paths = []
        for root, _, files in os.walk(Path('C:/Program Files/')):
            if 'Unity.exe' in files:
                possible_paths.append(Path(root) / 'Unity.exe')
        if not len(possible_paths):
            raise FileNotFoundError('No editor could not be found!')

        #check each path for the valid one
        for path in possible_paths:
            unity_version = run_cmd([path, '--version'], return_line=0)
            if editor_version in unity_version:
                unity_path = path
                break
        else:
            raise FileNotFoundError('Could not find valid version of Unity editor. Expected {}, found:{}'.format(editor_version, '\n'.join(possible_paths)))

        #update the editor path in the settings
        print('Updating path in settings file.')
        hook_settings['editor_path'] = str(unity_path)
        with open(settings_path, 'w') as f:
            updated_settings = json.dumps(hook_settings, indent=4)
            f.write(updated_settings)

    #create path variables
    print(Path.cwd())
    xml_path = project_path / 'Hooks/test_artifacts/results.xml'

    #run the test suite!
    print('Running tests.', flush=True)
    test_args = [unity_path, '-projectPath', project_path, '-runTests', '-testPlatform', target_platform, '-testResults', xml_path]
    if batchmode: test_args.insert(3, '-batchmode')
    unity_output = run_cmd(test_args, delim='')

    if not len(unity_output) == 0:
        #handle test runner failure
        if 'Multiple Unity instances cannot open the same project' in unity_output:
            tkinter.messagebox.showwarning('Failed to run tests!', 'Could not run tests because the Unity project is already open.' + 
                                           'If you\'d like to run the tests, please close the editor and try again.')
        else: 
            error_msg = f'Tests could not be run for the following reason:\n{unity_output}'
            tkinter.messagebox.showwarning('Failed to run tests!', error_msg)
    else:    
        #parse the xml -- navigate down to the xml tree to test fixtures and individual test cases
        print('Processing and formatting results.')

        parse_tree = ET.parse(xml_path).getroot()
        test_results = {} 
        process_results(parse_tree, test_results)

        if len(test_results) == 0:
            #empty results; tests failed for some reason
            if args.show_ui: tkinter.messagebox.showwarning('No tests were run!', 'Tests were run successfully, but no tests were run. Make sure that test suites exist.')
            else:            print('Tests were run successfully, but no tests were run. Make sure that test suites exist.')
            exit()   
        
        #make the file name
        results_name = f'results_{args.head_rev}.xlsx' #DO NOT REMOVE THE FILE EXTENSION FROM HERE
        print(f'Writing results to {results_name}')

        #format the results
        format_results(test_results, xml_path, results_name)

        #find failed tests
        failed_tests = [ f'\t{test_name}' for test_name, attrib in test_results.items() if attrib['result'] == 'Failed' ]
        if len(failed_tests) > 0:
            failed_tests_str = '\n'.join(failed_tests) 
            if args.show_ui:
                #no console; also, don't stop commit
                if not tkinter.messagebox.askokcancel('Warning!', f'The following tests have failed:\n{failed_tests_str}'):
                    exit()
            else:
                print(f'Failed tests:\n{failed_tests_str}', flush=True)

    if args.show_ui:
        tkinter.messagebox.showinfo('Done!', 'Tests complete.')
    else:
        print('Done!')

#steps up from cwd until it finds a unity project folder
def find_project_path() -> Path:
    start_dir = Path.cwd()
    MAX_SEARCH_DEPTH = 20
    unity_folders = ['Assets', 'Packages', 'ProjectSettings']
    for _ in range(MAX_SEARCH_DEPTH):
        #check that the current directory has all of the folders that a unity project needs
        dirs_in_cwd = [ dir for dir in os.listdir() if os.path.isdir(dir) ]
        for folder in unity_folders:
            if folder not in dirs_in_cwd:
                break
        else:
            #unity project found; restore the path and return the directory that it's in
            project_path = Path.cwd()
            os.chdir(start_dir)
            return project_path
        
        #step up
        os.chdir(Path.cwd().parent)
    else:
        raise RuntimeError('Could not find unity directory! Started in {}; ended in {}'.format(start_dir, Path.cwd()))

if __name__ == '__main__':
    #hide the mainapp window from tk 
    root = tkinter.Tk()
    root.attributes('-topmost',1)
    root.withdraw()

    #parse arguments
    parser = argparse.ArgumentParser(description='Test pipeline. Runs tests through Unity, then formats and uploads the results')
    parser.add_argument('--head_rev', type=int,
        help='The current head revision of the svn.')
    parser.add_argument('--show_ui', action='store_true',
        help='If included, will show message boxes.')
    args = parser.parse_args()
    print(f'head_rev: {args.head_rev}')
    print(f'show_ui: {args.show_ui}')

    #ensure that the current directory has 360
    print('Finding project path...', flush=True)
    project_path = find_project_path()
    print(f'Found project path: {project_path}')

    #check the test settings
    print('Reading settings.')
    settings_path = project_path / 'Hooks/hook_settings.json'
    with open(settings_path, 'r') as f:
        hook_settings = json.load(f)
        batchmode = hook_settings['batchmode']
        editor_path = hook_settings['editor_path']
        editor_version = hook_settings['editor_version']
        target_platform = hook_settings['target_platform']

    try:
        check_svn_clt()
        run_tests()
    except Exception as e:
        #always write errors to hooks folder
        hook_path = project_path / 'Hooks'
        os.chdir(hook_path)
        with open('error_log.txt', 'w') as f:
            error_msg = '{}\n{}'.format(str(datetime.now()), format_exc())
            f.write(error_msg)
            print(error_msg)
            tkinter.messagebox.showerror('Error!', f'The following exception occurred:\n{error_msg}')
            